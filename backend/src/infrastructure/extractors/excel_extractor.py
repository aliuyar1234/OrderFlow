"""Excel extractor - Rule-based extraction from Excel files (.xlsx, .xls).

Extracts order header and line items from Excel files using openpyxl.
Implements ExtractorPort interface for hexagonal architecture.

SSOT Reference: §7 (Extraction Logic)
"""

import io
import logging
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from typing import Optional, List, Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from domain.extraction.canonical_output import (
    CanonicalExtractionOutput,
    ExtractionOrderHeader,
    ExtractionLineItem,
)
from domain.extraction.confidence import calculate_confidence
from domain.extraction.ports import ExtractorPort, ExtractionResult
from domain.documents.ports.object_storage_port import ObjectStoragePort

logger = logging.getLogger(__name__)


class ExcelExtractor(ExtractorPort):
    """Excel file extractor using rule-based parsing.

    Extracts order data from .xlsx and .xls files using openpyxl library.
    Handles header row detection, decimal parsing, and column mapping.

    Features:
    - Automatic header row detection (looks for keywords like 'SKU', 'Qty', 'Artikel')
    - European decimal format support (comma as decimal separator)
    - Multi-sheet support (uses first/active sheet)
    - Whitespace normalization
    """

    def __init__(self, storage: ObjectStoragePort):
        """Initialize Excel extractor.

        Args:
            storage: Object storage adapter for retrieving files
        """
        self.storage = storage

    async def extract(self, document: Any) -> ExtractionResult:
        """Extract structured order data from Excel file.

        Args:
            document: Document entity with storage_key, mime_type, etc.

        Returns:
            ExtractionResult with canonical output or error
        """
        start_time = datetime.utcnow()
        metrics = {}

        try:
            # Retrieve file from storage
            logger.info(f"Retrieving Excel file from storage: {document.storage_key}")
            file_stream = await self.storage.retrieve_file(document.storage_key)

            # Read into BytesIO for openpyxl
            file_content = io.BytesIO(file_stream.read())
            file_stream.close()

            # Load workbook (read-only mode for performance)
            wb = openpyxl.load_workbook(file_content, read_only=True, data_only=True)
            sheet = wb.active  # Use active sheet (typically first sheet)

            if sheet is None:
                raise ValueError("Excel file has no active worksheet")

            sheet_name = sheet.title
            logger.info(f"Processing Excel sheet: {sheet_name}")

            # Detect header row
            header_row_idx = self._detect_header_row(sheet)
            logger.debug(f"Header row detected at index: {header_row_idx}")

            # Extract lines
            lines = []
            row_count = 0
            for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row_idx + 1), start=1):
                row_count += 1
                line = self._extract_line(row, row_idx)
                if line and self._is_valid_line(line):
                    lines.append(line)

            logger.info(f"Extracted {len(lines)} lines from {row_count} rows")

            # Extract header (look for order metadata in first rows)
            header = self._extract_header(sheet, header_row_idx)

            # Build canonical output
            output = CanonicalExtractionOutput(
                order=header,
                lines=lines,
                metadata={
                    'sheet_name': sheet_name,
                    'total_rows': row_count,
                    'header_row_idx': header_row_idx,
                }
            )

            # Calculate confidence
            confidence, confidence_breakdown = calculate_confidence(output)

            # Calculate metrics
            runtime_ms = int((datetime.utcnow() - start_time).total_seconds() * 1000)
            metrics = {
                'runtime_ms': runtime_ms,
                'row_count': row_count,
                'lines_extracted': len(lines),
                'sheet_name': sheet_name,
                'confidence': confidence,
                **confidence_breakdown,
            }

            logger.info(
                f"Excel extraction succeeded: {len(lines)} lines, "
                f"confidence={confidence:.3f}, runtime={runtime_ms}ms"
            )

            return ExtractionResult(
                success=True,
                output=output,
                confidence=confidence,
                metrics=metrics,
                extractor_version=self.version,
            )

        except Exception as e:
            runtime_ms = int((datetime.utcnow() - start_time).total_seconds() * 1000)
            logger.error(f"Excel extraction failed: {e}", exc_info=True)

            return ExtractionResult(
                success=False,
                error=str(e),
                confidence=0.0,
                metrics={'runtime_ms': runtime_ms, 'error_type': type(e).__name__},
                extractor_version=self.version,
            )

    def _detect_header_row(self, sheet: Worksheet) -> int:
        """Detect which row contains column headers.

        Looks for keywords like 'SKU', 'Qty', 'Artikel', 'Menge' in first 10 rows.

        Args:
            sheet: Excel worksheet

        Returns:
            Row index (1-based) containing headers, defaults to 1 if not found
        """
        header_keywords = [
            'sku', 'artikel', 'artikelnummer', 'product', 'item',
            'qty', 'menge', 'quantity', 'anzahl',
            'description', 'bezeichnung', 'beschreibung',
            'preis', 'price', 'unit price',
        ]

        for idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), start=1):
            if not row:
                continue

            # Convert row values to lowercase strings
            row_values = [
                str(cell).lower().strip()
                for cell in row
                if cell is not None
            ]

            # Check if any header keyword is in this row
            matches = sum(
                1 for keyword in header_keywords
                if any(keyword in value for value in row_values)
            )

            # If we find 2+ keywords, this is likely the header row
            if matches >= 2:
                return idx

        # Default to row 1 if no clear header found
        return 1

    def _extract_line(self, row, line_no: int) -> Optional[ExtractionLineItem]:
        """Extract line item from Excel row.

        Assumes column order: SKU, Description, Qty, UoM, Unit Price
        (This is a simplified heuristic - future versions could use column detection)

        Args:
            row: Excel row cells
            line_no: Line number (1-based)

        Returns:
            ExtractionLineItem or None if row is empty
        """
        cells = [cell.value for cell in row]

        # Skip empty rows
        if not cells or all(c is None or str(c).strip() == '' for c in cells):
            return None

        # Extract values with safe indexing
        customer_sku = self._get_cell_string(cells, 0)
        description = self._get_cell_string(cells, 1)
        qty = self._parse_decimal(self._get_cell_value(cells, 2))
        uom = self._get_cell_string(cells, 3)
        unit_price = self._parse_decimal(self._get_cell_value(cells, 4))
        currency = self._get_cell_string(cells, 5)

        return ExtractionLineItem(
            line_no=line_no,
            customer_sku=customer_sku,
            description=description,
            qty=qty,
            uom=uom,
            unit_price=unit_price,
            currency=currency,
        )

    def _get_cell_value(self, cells: List, index: int) -> Any:
        """Safely get cell value by index.

        Args:
            cells: List of cell values
            index: Cell index

        Returns:
            Cell value or None if index out of range
        """
        return cells[index] if index < len(cells) else None

    def _get_cell_string(self, cells: List, index: int) -> Optional[str]:
        """Safely get cell value as string.

        Args:
            cells: List of cell values
            index: Cell index

        Returns:
            String value or None if empty/missing
        """
        value = self._get_cell_value(cells, index)
        if value is None:
            return None
        str_value = str(value).strip()
        return str_value if str_value else None

    def _parse_decimal(self, value: Any) -> Optional[Decimal]:
        """Parse decimal value, handling European format (comma as decimal separator).

        Args:
            value: Value to parse (int, float, str, or None)

        Returns:
            Decimal or None if parsing fails
        """
        if value is None:
            return None

        # Handle numeric types
        if isinstance(value, (int, float)):
            try:
                return Decimal(str(value))
            except InvalidOperation:
                return None

        # Handle string type
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None

            # Handle European format: replace comma with dot
            # Remove thousand separators (spaces, dots in European format)
            value = value.replace(' ', '')  # Remove spaces
            value = value.replace('\u202f', '')  # Remove narrow no-break space

            # Check if this uses European format (comma for decimal)
            if ',' in value and '.' not in value:
                # Simple case: only comma (decimal separator)
                value = value.replace(',', '.')
            elif ',' in value and '.' in value:
                # Complex case: both comma and dot
                # Determine which is decimal separator by position
                comma_pos = value.rindex(',')
                dot_pos = value.rindex('.')

                if comma_pos > dot_pos:
                    # Comma is decimal separator (European: 1.234,56)
                    value = value.replace('.', '').replace(',', '.')
                else:
                    # Dot is decimal separator (US: 1,234.56)
                    value = value.replace(',', '')

            try:
                return Decimal(value)
            except (InvalidOperation, ValueError):
                logger.warning(f"Failed to parse decimal value: {value}")
                return None

        return None

    def _is_valid_line(self, line: ExtractionLineItem) -> bool:
        """Check if line has minimum required data.

        A valid line should have at least a SKU or description.

        Args:
            line: Line item to validate

        Returns:
            True if line has minimal required data
        """
        return line.customer_sku is not None or line.description is not None

    def _extract_header(self, sheet: Worksheet, header_row_idx: int) -> ExtractionOrderHeader:
        """Extract order header information from first rows.

        Searches for patterns like "Order Number: 12345" or "Date: 2024-01-15"
        in rows before the header row.

        Args:
            sheet: Excel worksheet
            header_row_idx: Index of header row (data starts after this)

        Returns:
            ExtractionOrderHeader with any found fields
        """
        header = ExtractionOrderHeader()

        # Search first N rows for header metadata
        for row in sheet.iter_rows(max_row=min(header_row_idx, 20), values_only=True):
            if not row:
                continue

            # Convert row to string for pattern matching
            row_text = ' '.join(str(cell) for cell in row if cell is not None).lower()

            # Try to extract order number
            if header.order_number is None:
                if 'order' in row_text or 'po' in row_text or 'bestellung' in row_text:
                    # Look for number in adjacent cells
                    for i, cell in enumerate(row):
                        if cell is None:
                            continue
                        cell_lower = str(cell).lower()
                        if 'order' in cell_lower or 'po' in cell_lower or 'bestellung' in cell_lower:
                            # Check next cell for number
                            if i + 1 < len(row) and row[i + 1] is not None:
                                header.order_number = str(row[i + 1]).strip()
                                break

            # Try to extract order date
            if header.order_date is None:
                for cell in row:
                    if isinstance(cell, date):
                        header.order_date = cell
                        break

        return header

    def supports(self, mime_type: str) -> bool:
        """Check if this extractor supports the given MIME type.

        Args:
            mime_type: MIME type string

        Returns:
            True if this extractor can handle Excel files
        """
        return mime_type in [
            'application/vnd.ms-excel',  # .xls
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
        ]

    @property
    def version(self) -> str:
        """Extractor version identifier.

        Returns:
            Version string for tracking
        """
        return 'excel_v1'

    @property
    def priority(self) -> int:
        """Priority for extractor selection (lower = higher priority).

        Returns:
            Priority (10 = high priority for rule-based extractors)
        """
        return 10
